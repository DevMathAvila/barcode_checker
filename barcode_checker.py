"""
Barcode Checker — Comparador de códigos de barras com Excel
Requisitos: pip install openpyxl
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
import datetime
import os
from openpyxl.utils import get_column_letter

# ─── Paleta de cores ────────────────────────────────────────────────────────
BG        = "#F8F8F6"
SURFACE   = "#FFFFFF"
BORDER    = "#DDDDD8"
TEXT      = "#1A1A18"
TEXT_MUTE = "#6B6B68"
ACCENT    = "#185FA5"
SUCCESS   = "#3B6D11"
DANGER    = "#A32D2D"
WARN_BG   = "#FAEEDA"

FONT      = ("Segoe UI", 10)
FONT_BOLD = ("Segoe UI", 10, "bold")
FONT_SM   = ("Segoe UI", 9)
FONT_LG   = ("Segoe UI", 13, "bold")
FONT_MONO = ("Consolas", 10)

TRACKING_FILE_NAME = "RUMO.Atualizacoes.xlsx"
ACTIVE_SHEET_NAME = "RUMO.Numeros_Ativos"
NOTFOUND_SHEET_NAME = "RUMO.Chips_Nofound"
BARCODE_COLUMN_INDEX = 10
SEARCH_BUTTON_TEXT = "▶  Iniciar comparação"
SEARCHING_BUTTON_TEXT = "Buscando…"


class BarcodeChecker(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Barcode Checker")
        self.geometry("820x620")
        self.minsize(700, 520)
        self.configure(bg=BG)
        self.resizable(True, True)

        self.excel_path = tk.StringVar(value="")
        self.sync_results_var = tk.BooleanVar(value=False)
        self.barcodes: list[str] = []
        self.results: list[dict] = []

        self._build_ui()
        self._bind_shortcuts()

    # ─── UI ──────────────────────────────────────────────────────────────────

    def _build_ui(self):
        # Cabeçalho
        header = tk.Frame(self, bg=ACCENT, pady=10)
        header.pack(fill="x")
        tk.Label(header, text="Barcode Checker",
                 font=FONT_LG, bg=ACCENT, fg="white").pack(side="left", padx=16)
        tk.Label(header, text="Comparador de códigos de barras com Excel",
                 font=FONT_SM, bg=ACCENT, fg="#B5D4F4").pack(side="left", padx=4)

        # Corpo principal
        body = tk.Frame(self, bg=BG, padx=14, pady=12)
        body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=2)
        body.rowconfigure(1, weight=1)

        # ── Coluna esquerda ──────────────────────────────────────────
        left = tk.Frame(body, bg=BG)
        left.grid(row=0, column=0, rowspan=2, sticky="nsew", padx=(0, 8))
        left.rowconfigure(1, weight=1)

        tk.Label(left, text="Códigos de barras", font=FONT_BOLD,
                 bg=BG, fg=TEXT).grid(row=0, column=0, sticky="w", pady=(0, 4))

        # Campo de entrada + botão adicionar
        entry_row = tk.Frame(left, bg=BG)
        entry_row.grid(row=1, column=0, sticky="ew")
        left.columnconfigure(0, weight=1)

        self.entry_var = tk.StringVar()
        self.entry = tk.Entry(entry_row, textvariable=self.entry_var,
                              font=FONT_MONO, relief="flat",
                              bg=SURFACE, fg=TEXT, insertbackground=TEXT,
                              highlightthickness=1, highlightbackground=BORDER,
                              highlightcolor=ACCENT, bd=0)
        self.entry.pack(side="left", fill="x", expand=True, ipady=6, padx=(0, 6))
        self.entry.bind("<Return>", lambda e: self._add_barcode())
        self.entry.focus_set()

        btn_add = tk.Button(entry_row, text="+ Adicionar",
                            font=FONT_SM, command=self._add_barcode,
                            bg=ACCENT, fg="white", relief="flat",
                            activebackground="#0C447C", activeforeground="white",
                            cursor="hand2", padx=10, pady=5)
        btn_add.pack(side="right")

        # Lista de códigos
        list_frame = tk.Frame(left, bg=SURFACE, relief="flat",
                              highlightthickness=1, highlightbackground=BORDER)
        list_frame.grid(row=2, column=0, sticky="nsew", pady=(6, 0))
        left.rowconfigure(2, weight=1)

        scrollbar = tk.Scrollbar(list_frame, orient="vertical")
        self.listbox = tk.Listbox(list_frame, font=FONT_MONO,
                                  bg=SURFACE, fg=TEXT, relief="flat",
                                  selectbackground=ACCENT, selectforeground="white",
                                  activestyle="none", bd=0,
                                  yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.listbox.yview)
        scrollbar.pack(side="right", fill="y")
        self.listbox.pack(fill="both", expand=True, padx=2, pady=2)

        # Botões lista
        btn_row = tk.Frame(left, bg=BG)
        btn_row.grid(row=3, column=0, sticky="ew", pady=(6, 0))

        tk.Button(btn_row, text="✎ Editar",
                  font=FONT_SM, command=self._edit_selected,
                  bg=SURFACE, fg=TEXT, relief="flat",
                  highlightthickness=1, highlightbackground=BORDER,
                  activebackground=BG, cursor="hand2",
                  padx=8, pady=4).pack(side="left", padx=(0, 6))

        tk.Button(btn_row, text="✕ Remover",
                  font=FONT_SM, command=self._remove_selected,
                  bg=SURFACE, fg=DANGER, relief="flat",
                  highlightthickness=1, highlightbackground=BORDER,
                  activebackground=BG, cursor="hand2",
                  padx=8, pady=4).pack(side="left", padx=(0, 6))

        tk.Button(btn_row, text="Limpar tudo",
                  font=FONT_SM, command=self._clear_all,
                  bg=SURFACE, fg=TEXT_MUTE, relief="flat",
                  highlightthickness=1, highlightbackground=BORDER,
                  activebackground=BG, cursor="hand2",
                  padx=8, pady=4).pack(side="left")

        self.count_label = tk.Label(btn_row, text="0 código(s)",
                                    font=FONT_SM, bg=BG, fg=TEXT_MUTE)
        self.count_label.pack(side="right")

        # ── Coluna direita ───────────────────────────────────────────
        right = tk.Frame(body, bg=BG)
        right.grid(row=0, column=1, rowspan=2, sticky="nsew")
        right.rowconfigure(6, weight=1)
        right.columnconfigure(0, weight=1)

        # Seleção do Excel
        tk.Label(right, text="Arquivo Excel", font=FONT_BOLD,
                 bg=BG, fg=TEXT).grid(row=0, column=0, sticky="w", pady=(0, 4))

        excel_row = tk.Frame(right, bg=BG)
        excel_row.grid(row=1, column=0, sticky="ew", pady=(16, 0))

        self.path_label = tk.Label(excel_row, textvariable=self.excel_path,
                                   font=FONT_SM, bg=SURFACE, fg=TEXT_MUTE,
                                   anchor="w", width=36, padx=8,
                                   relief="flat",
                                   highlightthickness=1,
                                   highlightbackground=BORDER)
        self.path_label.pack(side="left", fill="x", expand=True, ipady=5, padx=(0, 6))

        tk.Button(excel_row, text="Selecionar Excel",
                  font=FONT_SM, command=self._pick_excel,
                  bg=SURFACE, fg=TEXT, relief="flat",
                  highlightthickness=1, highlightbackground=BORDER,
                  activebackground=BG, cursor="hand2",
                  padx=10, pady=5).pack(side="right")

        self.sync_check = tk.Checkbutton(
            right,
            text="Atualizar planilha RUMO",
            variable=self.sync_results_var,
            font=FONT_SM,
            bg=BG,
            fg=TEXT,
            activebackground=BG,
            activeforeground=TEXT,
            selectcolor=SURFACE,
            cursor="hand2",
            anchor="w",
        )
        self.sync_check.grid(row=2, column=0, sticky="w", pady=(10, 0))

        # Botão buscar
        self.btn_search = tk.Button(right, text=SEARCH_BUTTON_TEXT,
                                    font=FONT_BOLD, command=self._run_search,
                                    bg=SUCCESS, fg="white", relief="flat",
                                    activebackground="#27500A",
                                    activeforeground="white",
                                    cursor="hand2", padx=16, pady=8)
        self.btn_search.grid(row=3, column=0, sticky="e", pady=(12, 0))

        # Separador
        ttk.Separator(right, orient="horizontal").grid(
            row=4, column=0, sticky="ew", pady=(10, 6))

        # Resultado / Histórico
        tk.Label(right, text="Histórico de resultados", font=FONT_BOLD,
                 bg=BG, fg=TEXT).grid(row=5, column=0, sticky="w", pady=(0, 4))

        # Treeview resultado
        tree_frame = tk.Frame(right, bg=SURFACE,
                              highlightthickness=1, highlightbackground=BORDER)
        tree_frame.grid(row=6, column=0, sticky="nsew")

        cols = ("barcode", "sheet", "col", "row", "cell", "status")
        self.tree = ttk.Treeview(tree_frame, columns=cols,
                                 show="headings", height=14)
        self._style_tree()

        headers = {
            "barcode": ("Código", 160),
            "sheet":   ("Aba",     90),
            "col":     ("Coluna",  60),
            "row":     ("Linha",   55),
            "cell":    ("Célula",  65),
            "status":  ("Status", 220),
        }
        for col, (label, width) in headers.items():
            self.tree.heading(col, text=label)
            self.tree.column(col, width=width, anchor="center",
                             minwidth=40, stretch=(col == "barcode"))

        vsb = tk.Scrollbar(tree_frame, orient="vertical",
                           command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)

        # Rodapé
        footer = tk.Frame(right, bg=BG)
        footer.grid(row=7, column=0, sticky="ew", pady=(6, 0))

        tk.Button(footer, text="Exportar histórico (.txt)",
                  font=FONT_SM, command=self._export_history,
                  bg=SURFACE, fg=TEXT, relief="flat",
                  highlightthickness=1, highlightbackground=BORDER,
                  activebackground=BG, cursor="hand2",
                  padx=10, pady=4).pack(side="left")

        tk.Button(footer, text="Limpar histórico",
                  font=FONT_SM, command=self._clear_results,
                  bg=SURFACE, fg=TEXT_MUTE, relief="flat",
                  highlightthickness=1, highlightbackground=BORDER,
                  activebackground=BG, cursor="hand2",
                  padx=10, pady=4).pack(side="left", padx=6)

        self.status_bar = tk.Label(footer, text="Pronto.",
                                   font=FONT_SM, bg=BG, fg=TEXT_MUTE, anchor="e")
        self.status_bar.pack(side="right")

    def _style_tree(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("Treeview",
                        background=SURFACE, foreground=TEXT,
                        fieldbackground=SURFACE,
                        rowheight=24, font=FONT_SM, borderwidth=0)
        style.configure("Treeview.Heading",
                        background=BG, foreground=TEXT_MUTE,
                        font=FONT_SM, relief="flat")
        style.map("Treeview",
                  background=[("selected", ACCENT)],
                  foreground=[("selected", "white")])
        self.tree.tag_configure("found",   foreground=SUCCESS, font=FONT_SM)
        self.tree.tag_configure("notfound", foreground=DANGER,  font=FONT_SM)

    def _bind_shortcuts(self):
        self.listbox.bind("<Delete>", lambda e: self._remove_selected())
        self.entry.bind("<Escape>",   lambda e: self.entry_var.set(""))

    # ─── Lógica de lista ─────────────────────────────────────────────────────

    def _add_barcode(self):
        code = self.entry_var.get().strip()
        if not code:
            return
        if code in self.barcodes:
            messagebox.showwarning("Duplicado",
                                   f"O código '{code}' já está na lista.")
            return
        self.barcodes.append(code)
        self.listbox.insert("end", code)
        self.entry_var.set("")
        self._update_count()
        self.entry.focus_set()

    def _remove_selected(self):
        sel = self.listbox.curselection()
        if not sel:
            messagebox.showinfo("Atenção", "Selecione um código para remover.")
            return
        idx = sel[0]
        self.barcodes.pop(idx)
        self.listbox.delete(idx)
        self._update_count()

    def _edit_selected(self):
        sel = self.listbox.curselection()
        if not sel:
            messagebox.showinfo("Atenção", "Selecione um código para editar.")
            return
        idx = sel[0]
        current = self.barcodes[idx]
        win = tk.Toplevel(self)
        win.title("Editar código")
        win.geometry("340x110")
        win.resizable(False, False)
        win.configure(bg=BG)
        win.grab_set()

        tk.Label(win, text="Novo valor:", font=FONT, bg=BG, fg=TEXT
                 ).pack(padx=16, pady=(16, 4), anchor="w")

        var = tk.StringVar(value=current)
        e = tk.Entry(win, textvariable=var, font=FONT_MONO,
                     relief="flat", bg=SURFACE, fg=TEXT,
                     insertbackground=TEXT,
                     highlightthickness=1, highlightbackground=BORDER,
                     highlightcolor=ACCENT, bd=0)
        e.pack(fill="x", padx=16, ipady=6)
        e.icursor("end")
        e.focus_set()

        def save(event=None):
            new = var.get().strip()
            if not new:
                return
            if new != current and new in self.barcodes:
                messagebox.showwarning("Duplicado",
                                       f"'{new}' já existe na lista.", parent=win)
                return
            self.barcodes[idx] = new
            self.listbox.delete(idx)
            self.listbox.insert(idx, new)
            win.destroy()

        e.bind("<Return>", save)
        tk.Button(win, text="Salvar", font=FONT_SM, command=save,
                  bg=ACCENT, fg="white", relief="flat",
                  activebackground="#0C447C",
                  cursor="hand2", padx=12, pady=4
                  ).pack(side="right", padx=16, pady=(8, 0))

    def _clear_all(self):
        if self.barcodes and messagebox.askyesno(
                "Confirmar", "Limpar todos os códigos da lista?"):
            self.barcodes.clear()
            self.listbox.delete(0, "end")
            self._update_count()

    def _update_count(self):
        n = len(self.barcodes)
        self.count_label.config(text=f"{n} código(s)")

    # ─── Excel ───────────────────────────────────────────────────────────────

    def _pick_excel(self):
        path = filedialog.askopenfilename(
            title="Selecionar arquivo Excel",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls"), ("Todos", "*.*")])
        if path:
            self.excel_path.set(path)
            self._set_status(f"Excel: {os.path.basename(path)}")

    def _get_tracking_workbook_path(self, source_path: str) -> str:
        folder = os.path.dirname(source_path) or os.getcwd()
        return os.path.join(folder, TRACKING_FILE_NAME)

    def _get_source_headers(self, workbook) -> list:
        if not workbook.sheetnames:
            return [""] * BARCODE_COLUMN_INDEX

        header_row = [cell.value for cell in workbook[workbook.sheetnames[0]][1]]
        while header_row and (header_row[-1] is None or str(header_row[-1]).strip() == ""):
            header_row.pop()

        headers = ["" if value is None else value for value in header_row]
        if len(headers) < BARCODE_COLUMN_INDEX:
            headers.extend([""] * (BARCODE_COLUMN_INDEX - len(headers)))
        return headers

    def _ensure_sheet_headers(self, ws, headers: list):
        existing = [ws.cell(row=1, column=col).value for col in range(1, len(headers) + 1)]
        if any(value not in (None, "") for value in existing):
            return

        for col_idx, value in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=value)

    def _open_tracking_workbook(self, source_wb, source_path: str):
        tracking_path = self._get_tracking_workbook_path(source_path)
        headers = self._get_source_headers(source_wb)
        created = not os.path.exists(tracking_path)

        if created:
            tracking_wb = openpyxl.Workbook()
            default_sheet = tracking_wb.active
            tracking_wb.remove(default_sheet)
        else:
            tracking_wb = openpyxl.load_workbook(tracking_path)

        for sheet_name in (ACTIVE_SHEET_NAME, NOTFOUND_SHEET_NAME):
            if sheet_name not in tracking_wb.sheetnames:
                tracking_wb.create_sheet(sheet_name)
            self._ensure_sheet_headers(tracking_wb[sheet_name], headers)

        return (
            tracking_wb,
            tracking_wb[ACTIVE_SHEET_NAME],
            tracking_wb[NOTFOUND_SHEET_NAME],
            tracking_path,
            headers,
            created,
        )

    def _row_values_until_last_filled(self, ws, row_number: int) -> list:
        row_values = [cell.value for cell in ws[row_number]]
        while row_values and (row_values[-1] is None or str(row_values[-1]).strip() == ""):
            row_values.pop()

        if len(row_values) < BARCODE_COLUMN_INDEX:
            row_values.extend([""] * (BARCODE_COLUMN_INDEX - len(row_values)))
        return row_values

    def _sheet_has_barcode(self, ws, code: str) -> bool:
        if ws.max_row < 2:
            return False

        for (value,) in ws.iter_rows(
                min_row=2,
                max_row=ws.max_row,
                min_col=BARCODE_COLUMN_INDEX,
                max_col=BARCODE_COLUMN_INDEX,
                values_only=True):
            if value is not None and str(value).strip() == code:
                return True
        return False

    def _append_found_to_tracking(self, ws, code: str, row_values: list) -> bool:
        if self._sheet_has_barcode(ws, code):
            return False

        values = list(row_values)
        if len(values) < BARCODE_COLUMN_INDEX:
            values.extend([""] * (BARCODE_COLUMN_INDEX - len(values)))
        ws.append(values)
        return True

    def _append_notfound_to_tracking(self, ws, code: str, header_len: int) -> bool:
        if self._sheet_has_barcode(ws, code):
            return False

        row_values = [""] * max(header_len, BARCODE_COLUMN_INDEX)
        row_values[BARCODE_COLUMN_INDEX - 1] = code
        ws.append(row_values)
        return True

    # ─── Comparação ──────────────────────────────────────────────────────────

    def _run_search(self):
        if not self.barcodes:
            messagebox.showwarning("Atenção", "Adicione ao menos um código antes de buscar.")
            return
        path = self.excel_path.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning("Atenção", "Selecione um arquivo Excel válido.")
            return

        self.results.clear()
        for item in self.tree.get_children():
            self.tree.delete(item)

        self.btn_search.config(state="disabled", text=SEARCHING_BUTTON_TEXT)
        self.update_idletasks()

        wb = None
        tracking_wb = None
        active_ws = None
        notfound_ws = None
        tracking_path = ""
        tracking_headers: list = []
        tracking_created = False
        sync_enabled = self.sync_results_var.get()

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as exc:
            messagebox.showerror("Erro ao abrir Excel", str(exc))
            return

        try:
            if sync_enabled:
                tracking_wb, active_ws, notfound_ws, tracking_path, tracking_headers, tracking_created = (
                    self._open_tracking_workbook(wb, path)
                )

            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Índice rápido: valor → lista de (aba, col_letter, row_num)
            index: dict[str, list] = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            key = str(cell.value).strip()
                            if key not in index:
                                index[key] = []
                            col_letter = get_column_letter(cell.column)
                            index[key].append({
                                "sheet": sheet_name,
                                "col":   col_letter,
                                "row":   cell.row,
                                "cell":  f"{col_letter}{cell.row}",
                            })

            found_count = 0
            not_found_count = 0

            for code in self.barcodes:
                matches = index.get(code, [])
                sync_status = ""

                if matches and sync_enabled:
                    first_match = matches[0]
                    row_values = self._row_values_until_last_filled(
                        wb[first_match["sheet"]],
                        first_match["row"],
                    )
                    was_added = self._append_found_to_tracking(active_ws, code, row_values)
                    sync_status = "Adicionado" if was_added else "Já há - Duplicado"

                if not matches and sync_enabled:
                    was_added = self._append_notfound_to_tracking(
                        notfound_ws,
                        code,
                        len(tracking_headers),
                    )
                    sync_status = "Adicionado" if was_added else "Já há - Duplicado"

                if matches:
                    status_text = "Encontrado"
                    if sync_status:
                        status_text = f"{status_text} | {sync_status}"

                    for m in matches:
                        record = {
                            "ts":      timestamp,
                            "barcode": code,
                            "sheet":   m["sheet"],
                            "col":     m["col"],
                            "row":     m["row"],
                            "cell":    m["cell"],
                            "status":  status_text,
                        }
                        self.results.append(record)
                        self.tree.insert("", "end",
                                         values=(code, m["sheet"], m["col"],
                                                 m["row"], m["cell"], f"✔ {status_text}"),
                                         tags=("found",))
                    found_count += 1
                else:
                    status_text = "Não encontrado"
                    if sync_status:
                        status_text = f"{status_text} | {sync_status}"

                    record = {
                        "ts":      timestamp,
                        "barcode": code,
                        "sheet":   "—",
                        "col":     "—",
                        "row":     "—",
                        "cell":    "—",
                        "status":  status_text,
                    }
                    self.results.append(record)
                    self.tree.insert("", "end",
                                     values=(code, "—", "—", "—", "—",
                                             f"✘ {status_text}"),
                                     tags=("notfound",))
                    not_found_count += 1

            if tracking_wb is not None:
                tracking_wb.save(tracking_path)

            status_msg = (
                f"Concluído — {found_count} encontrado(s), "
                f"{not_found_count} não encontrado(s)."
            )
            if tracking_wb is not None:
                verb = "criada e atualizada" if tracking_created else "atualizada"
                status_msg += f" Planilha auxiliar {verb}: {os.path.basename(tracking_path)}."
            self._set_status(status_msg)
        except Exception as exc:
            messagebox.showerror("Erro na comparação", str(exc))
        finally:
            if tracking_wb is not None:
                tracking_wb.close()
            if wb is not None:
                wb.close()
            self.btn_search.config(state="normal", text=SEARCH_BUTTON_TEXT)

    # ─── Exportar ────────────────────────────────────────────────────────────

    def _export_history(self):
        if not self.results:
            messagebox.showinfo("Histórico vazio",
                                "Não há resultados para exportar.")
            return
        path = filedialog.asksaveasfilename(
            title="Salvar histórico",
            defaultextension=".txt",
            filetypes=[("Texto", "*.txt"), ("CSV", "*.csv"), ("Todos", "*.*")])
        if not path:
            return
        sep = "," if path.endswith(".csv") else "\t"
        header_line = sep.join(
            ["Timestamp", "Código", "Aba", "Coluna", "Linha", "Célula", "Status"])
        lines = [header_line]
        for r in self.results:
            lines.append(sep.join([
                r["ts"], r["barcode"], r["sheet"],
                r["col"], str(r["row"]), r["cell"], r["status"]
            ]))
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))
            messagebox.showinfo("Exportado",
                                f"Histórico salvo em:\n{path}")
        except Exception as exc:
            messagebox.showerror("Erro ao salvar", str(exc))

    def _clear_results(self):
        if self.results and messagebox.askyesno(
                "Confirmar", "Limpar todo o histórico de resultados?"):
            self.results.clear()
            for item in self.tree.get_children():
                self.tree.delete(item)
            self._set_status("Histórico limpo.")

    # ─── Utils ───────────────────────────────────────────────────────────────

    def _set_status(self, msg: str):
        self.status_bar.config(text=msg)


if __name__ == "__main__":
    app = BarcodeChecker()
    app.mainloop()
